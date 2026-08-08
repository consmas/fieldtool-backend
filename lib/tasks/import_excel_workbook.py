#!/usr/bin/env python3
import json
import os
import sys
from datetime import date, datetime

from openpyxl import load_workbook


def normalize_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        return None if text == "" else text
    return value


def is_header_row(row):
    values = [str(v or "").strip().lower() for v in row if v is not None]
    if not values:
        return False
    if any("truck registration number" in value for value in values):
        return False

    matches = sum(1 for value in values if any(token in value for token in ["date", "customer", "waybill", "destination", "no. of stops", "additional km", "fuel cost", "base fee", "additional fee", "total fee", "trip id", "driver"]))
    has_date = any("date" in value for value in values)
    has_financial = any(any(token in value for token in ["base fee", "additional fee", "total fee", "fuel cost"]) for value in values)
    has_customer = any("customer" in value for value in values)

    return matches >= 3 or (has_date and (has_customer or any(any(token in value for token in ["waybill", "destination"]) for value in values) or has_financial))


def normalize_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            return datetime.fromisoformat(text.replace(" ", "T")).date().isoformat()
        except ValueError:
            return text
    return str(value)


def should_skip_payload(payload):
    values = [str(v).strip() for v in payload.values() if v is not None and str(v).strip() != ""]
    if not values:
        return True

    joined = " ".join(values).lower()
    if any(token in joined for token in ["holiday", "weekend", "downtime", "total", "truck registration number", "not available", "not submitted"]):
        return True

    if payload.get("Date") in {"Date", "date"}:
        return True
    if payload.get("Customer Name") in {"Customer Name", "customer name"}:
        return True

    return False


def parse_book(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = wb.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))

    header_index = None
    for index, row in enumerate(rows):
        if is_header_row(row):
            header_index = index
            break

    if header_index is None:
        return []

    headers = [normalize_value(v) for v in rows[header_index]]
    payloads = []
    for row in rows[header_index + 1:]:
        if not any(v is not None and str(v).strip() != "" for v in row):
            continue
        if not row:
            continue
        if str(row[0]).strip().lower() in {"holiday", "weekend", "downtime", "total"}:
            continue
        payload = {}
        for idx, header in enumerate(headers):
            if header is None:
                continue
            payload[header] = normalize_value(row[idx]) if idx < len(row) else None

        if should_skip_payload(payload):
            continue

        payloads.append({
            "entry_date": normalize_date(payload.get("Date")) if payload.get("Date") else None,
            "waybill_no": payload.get("Waybill No.") or payload.get("Waybill Number") or payload.get("Waybill"),
            "truck_id": payload.get("Truck ID") or payload.get("Vehicle") or payload.get("Truck") or payload.get("Truck Registration Number"),
            "driver_name": payload.get("Driver Name") or payload.get("Driver"),
            "destination": payload.get("Destination") or payload.get("Drop Off Location") or payload.get("Dropoff Location"),
            "cargo_type": payload.get("Cargo Type") or payload.get("Cargo") or payload.get("Material Description"),
            "origin": payload.get("Origin") or payload.get("Pickup Location") or payload.get("Loading Point"),
            "expected_revenue": payload.get("Total Fee") or payload.get("Amount Due") or payload.get("Amount Invoiced") or payload.get("Base Fee"),
            "status": "Completed" if payload.get("Total Fee") else "Pending",
            "notes": payload.get("Customer Name") or payload.get("Remarks") or payload.get("Notes"),
            "details": {
                "customer_name": payload.get("Customer Name"),
                "base_fee": payload.get("Base Fee"),
                "additional_fee": payload.get("Additional Fee"),
                "additional_km_travelled": payload.get("Additional km Travelled"),
                "fuel_cost_per_litre": payload.get("Fuel Cost Per Litre*"),
                "no_of_stops": payload.get("No. of Stops")
            }
        })
    return payloads


if __name__ == "__main__":
    if len(sys.argv) < 2:
        raise SystemExit("Usage: import_excel_workbook.py <path>")
    path = sys.argv[1]
    print(json.dumps(parse_book(path)))
